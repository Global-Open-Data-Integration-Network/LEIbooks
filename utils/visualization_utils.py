import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Circle
import math
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from typing import Any, Optional, List
import ipywidgets as widgets
from IPython.display import display, Markdown, HTML

class Visualizations:
    """
    Creates visulizations for LEI data

    - Creates a stacked bar chart using matplotlib for better control over stacking.
    """

    def create_matplotlib_stacked_chart(by_mapping, level_1_subset, top_n=5):
        """
        Create a stacked bar chart using matplotlib for better control over stacking.
        """

        # Find the jurisdiction column name
        jurisdiction_col = None
        for col in level_1_subset.columns:
            if "LegalJurisdiction" in col:
                jurisdiction_col = col
                break

        if jurisdiction_col is None:
            raise ValueError("No LegalJurisdiction column found in level_1_subset")

        # Get top N jurisdictions by total mapping pairs
        jurisdiction_counts = {}

        # Merge each mapping with level_1 data and count by jurisdiction
        for mapping_name, mapping_df in by_mapping.items():
            merged = pd.merge(level_1_subset, mapping_df, on="LEI", how="inner")
            jurisdiction_counts[mapping_name] = merged[jurisdiction_col].value_counts()

        # Get top N jurisdictions
        all_jurisdictions = set()
        for counts in jurisdiction_counts.values():
            all_jurisdictions.update(counts.index)

        total_counts = {}
        for jurisdiction in all_jurisdictions:
            total = sum(
                counts.get(jurisdiction, 0) for counts in jurisdiction_counts.values()
            )
            total_counts[jurisdiction] = total

        top_jurisdictions = sorted(
            total_counts.items(), key=lambda x: x[1], reverse=True
        )[:top_n]
        top_jurisdiction_names = [j[0] for j in top_jurisdictions]

        # Prepare data for stacking
        mapping_types = list(by_mapping.keys())
        mapping_labels = [
            mapping.replace("-lei", "").upper() for mapping in mapping_types
        ]

        # Create data matrix for stacking
        data_matrix = []
        for jurisdiction in top_jurisdiction_names:
            row = []
            for mapping_name in mapping_types:
                count = jurisdiction_counts[mapping_name].get(jurisdiction, 0)
                row.append(count)
            data_matrix.append(row)

        data_matrix = np.array(data_matrix)

        # Create the plot with wider figure to accommodate wider bars
        fig, ax = plt.subplots(figsize=(14, 8))

        # Use turquoise color scheme consistent with existing visualizations
        colors = ["#123235", "#79D7C5", "#403E74", "#4D979B", "#4DA2F8"]

        # Create stacked bars
        bottom = np.zeros(len(top_jurisdiction_names))
        bar_width = 0.7

        for i, (mapping_type, label) in enumerate(zip(mapping_types, mapping_labels)):
            values = data_matrix[:, i]
            ax.bar(
                top_jurisdiction_names,
                values,
                bottom=bottom,
                label=label,
                color=colors[i % len(colors)],
                alpha=1.0,
                width=bar_width,
            )

            # Add value labels on significant segments
            for j, (jurisdiction, value) in enumerate(
                zip(top_jurisdiction_names, values)
            ):
                if (
                    value > 0 and value > max(data_matrix.flatten()) * 0.03
                ):  # Only show significant values
                    label_y = bottom[j] + value / 2
                    ax.text(
                        j,
                        label_y,
                        f"{value:,}",
                        ha="center",
                        va="center",
                        fontsize=9,
                        fontweight="bold",
                        color="white"
                        if value > max(data_matrix.flatten()) * 0.08
                        else "black",
                    )

            bottom += values

        # Customize the plot
        ax.set_title(
            "Distribution of Mapping Pairs by Jurisdiction and Mapping Type",
            fontsize=16,
            fontweight="bold",
            pad=20,
        )
        ax.set_xlabel("Legal Jurisdiction", fontsize=12, fontweight="bold")
        ax.set_ylabel("Number of Mapping Pairs", fontsize=12, fontweight="bold")

        # Rotate x-axis labels
        plt.xticks(rotation=45, ha="right")

        # Add some padding to prevent labels from touching borders
        ax.margins(x=0.05, y=0.05)

        # Format y-axis with commas
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{int(x):,}"))

        # Add legend with controlled order (reverse order to match stacking from bottom to top)
        handles, labels = ax.get_legend_handles_labels()
        # Reverse the order so legend matches stacking order (bottom to top)
        ax.legend(
            handles[::-1],
            labels[::-1],
            title="Mapping Type",
            bbox_to_anchor=(1.05, 1),
            loc="upper left",
        )

        # Add grid for better readability
        ax.grid(True, alpha=0.3, axis="y")

        # Use tight_layout with padding to ensure proper spacing
        plt.tight_layout(pad=2.0)
        plt.show()

        # Print detailed breakdown
        print("\nDetailed Breakdown by Jurisdiction:")
        print("=" * 60)
        for i, jurisdiction in enumerate(top_jurisdiction_names):
            print(f"\n{jurisdiction}:")
            total_for_jurisdiction = 0
            for j, (mapping_type, label) in enumerate(
                zip(mapping_types, mapping_labels)
            ):
                count = data_matrix[i, j]
                if count > 0:
                    percentage = (count / sum(data_matrix[i, :])) * 100
                    print(f"  {label}: {count:,} ({percentage:.1f}%)")
                    total_for_jurisdiction += count
            print(f"  Total: {total_for_jurisdiction:,}")

        return data_matrix, top_jurisdiction_names, mapping_labels

    def draw_bar_chart_jurisdictions(top20: list[tuple[str, int]], mapping_pairs: str):
        # Plot bar chart
        plt.figure(figsize=(6, 4))
        ax = top20.plot(kind="bar", color="Turquoise")
        plt.title(f"Top 20 Legal Jurisdictions for {mapping_pairs}")
        plt.xlabel("Legal Jurisdiction")
        plt.ylabel("Number of mapping pairs")
        plt.xticks(rotation=45)

        # Format y-axis labels with thousand separators
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f"{x:,.0f}"))

        plt.tight_layout()
        plt.show()

    def draw_star_map(
        center_label: str, center_value: int, data: list[tuple[str, int]]
    ):
        """
        Draw a simple hub-and-spoke map

        Args:
            center_label: Text label inside the center circle.
            center_value: Number inside the center circle.
            data: - a list of (label, value) tuples.
        """
        leaves = data

        n = len(leaves)
        if n == 0:
            raise ValueError("No leaves provided")

        # Plot setup
        fig, ax = plt.subplots(figsize=(7, 7))
        ax.set_aspect("equal")
        ax.axis("off")

        # Style defaults
        center_circle_color = "#509C9E"
        center_line_color = "#85D4D8"
        leaf_circle_color = "#79D7C5"
        line_color = "dimgray"
        halo_color = "cyan"
        center_radius = 1.5
        leaf_radius = 0.75
        halo_expand = 0.1
        ring_radius = 3.0

        # --- Center node ---
        # Halo (bigger, transparent)
        ax.add_patch(
            Circle(
                (0, 0),
                center_radius + halo_expand,
                facecolor=halo_color,
                alpha=0.2,
                edgecolor="none",
            )
        )
        ax.add_patch(
            Circle(
                (0, 0),
                center_radius,
                facecolor=center_circle_color,
                edgecolor=center_line_color,
                linewidth=2,
            )
        )
        ax.text(
            0,
            0,
            f"{center_label}\n{center_value:,}",
            ha="center",
            va="center",
            fontsize=12,
        )

        # --- Outer nodes ---
        angles = [2 * math.pi * i / n for i in range(n)]
        for (label, value), theta in zip(leaves, angles):
            x, y = ring_radius * math.cos(theta), ring_radius * math.sin(theta)

            # Direction vector
            dx, dy = x, y
            dist = math.hypot(dx, dy)
            dxn, dyn = dx / dist, dy / dist

            # Edge-to-edge line
            x_start, y_start = dxn * center_radius, dyn * center_radius
            x_end, y_end = x - dxn * leaf_radius, y - dyn * leaf_radius
            ax.plot(
                [x_start, x_end],
                [y_start, y_end],
                color=line_color,
                linewidth=1,
                linestyle="--",
            )

            # Halo
            ax.add_patch(
                Circle(
                    (x, y),
                    leaf_radius + halo_expand,
                    facecolor=halo_color,
                    alpha=0.2,
                    edgecolor="none",
                )
            )
            # Leaf circle
            ax.add_patch(
                Circle(
                    (x, y), leaf_radius, facecolor=leaf_circle_color, edgecolor="none"
                )
            )
            ax.text(x, y, f"{label}\n{value:,}", ha="center", va="center", fontsize=10)

        # Adjust limits
        pad = ring_radius + leaf_radius + 0.7
        ax.set_xlim(-pad, pad)
        ax.set_ylim(-pad, pad)

        ax.set_title(
            "Existing Mappings to LEI data", fontsize=12, fontweight="bold", pad=20
        )

        plt.show()

class LegalEntityEventsVisualizer:
    """
    Handles widgets + HTML display for Legal Entity Events.
    """

    def __init__(self, data_provider: Any, event_types: Optional[List[str]] = None) -> None:
        self.data = data_provider
        self.event_types = event_types  # e.g. ["CHANGE_LEGAL_NAME", "CHANGE_HQ_ADDRESS"]
        self.description_widget = widgets.HTML()
        self.lei_dropdown = widgets.Dropdown()
        self.date_dropdown = widgets.Dropdown()
        self.output_widget = widgets.Output()

        self.date_dropdown.style.description_width = "500px"
        self.date_dropdown.layout.width = "350px"

        

    # ---------- Internal viz helpers ----------

    def _style_and_display_table(self, df: pd.DataFrame) -> None:
        """
        Render the event details as a simple HTML table.
        """
        if df.empty:
            display(HTML("<em>No additional fields captured for this event.</em>"))
            return

        table_html = df.to_html(index=False, escape=False)
        display(
            HTML(
                """
                <style>
                    .lei-event-table table {
                        table-layout: fixed;
                        width: 100%;
                        border-collapse: collapse;
                        font-size: 14px;
                    }
                    .lei-event-table th,
                    .lei-event-table td {
                        text-align: center;
                        padding: 6px 10px;
                        border: 1px solid #ccc;
                        vertical-align: middle;

                        /* force long text to wrap */
                        white-space: normal;
                        word-wrap: break-word;
                        overflow-wrap: anywhere;
                    }
                    
                    .lei-event-table th {
                        background-color: #51DAC5;
                        font-weight: bold;
                    }
                </style>
                <div class="lei-event-table">
                """
                + table_html
                + "</div>"
            )
        )

    # ---------- Widget callbacks ----------

    def _update_dates(self, change) -> None:
        """Update the date dropdown when an LEI is selected."""
        if self.date_dropdown is None or self.output_widget is None:
            return

        selected_lei = change.get("new")

        # Clear old output
        with self.output_widget:
            self.output_widget.clear_output()

        df_lei = self.data.get_lei_dataframe(selected_lei)
        if df_lei.empty:
            self.date_dropdown.unobserve(self._update_table_multi, names="value")
            self.date_dropdown.options = []
            print("No modifications found for this LEI.")
            return

        valid_dates = self.data.get_valid_dates_for_lei(selected_lei)

        if not valid_dates:
            self.date_dropdown.unobserve(self._update_table_multi, names="value")
            self.date_dropdown.options = []
            print(f"No dates where '{self.data.required_field}' was modified.")
            return

        # Update the date dropdown safely
        self.date_dropdown.unobserve(self._update_table_multi, names="value")
        self.date_dropdown.options = valid_dates
        self.date_dropdown.value = valid_dates[0]
        self.date_dropdown.observe(self._update_table_multi, names="value")

        # Update table for default date
        self._update_table_multi({"new": valid_dates[0]})

    def _update_table_multi(self, change) -> None:
        """Display modification tables for selected date, grouped by event type."""
        if self.lei_dropdown is None or self.output_widget is None:
            return

        selected_lei = self.lei_dropdown.value
        selected_date = change.get("new")

        df_lei = self.data.get_lei_dataframe(selected_lei)

        with self.output_widget:
            self.output_widget.clear_output()

            if df_lei.empty:
                print("No data for this LEI.")
                return

            if (
                "updated_date" not in df_lei.columns
                or selected_date not in df_lei["updated_date"].unique()
            ):
                print("No data for this date.")
                return

            # Event tables from data provider – apply optional filter
            event_tables = self.data.get_event_tables_for_day(
                selected_lei,
                selected_date,
                event_types=self.event_types,
            )

            if not event_tables:
                # Distinguish between "no events at all" vs
                # "events exist but not of requested types"
                all_events = self.data.get_event_tables_for_day(
                    selected_lei,
                    selected_date,
                    event_types=None,  # no filter
                )

                if not all_events:
                    print(
                        f"No records with '{self.data.required_field}' "
                        f"for LEI {selected_lei} on {selected_date}."
                    )
                else:
                    # There are events, but none match the requested types
                    types_txt = ", ".join(self.event_types or [])
                    print(
                        f"No events of requested types [{types_txt}] "
                        f"for LEI {selected_lei} on {selected_date}."
                    )
                return


            # Event tables
            for event_label, visible_df in event_tables.items():
                display(Markdown(f"### Event: `{event_label}`"))
                self._style_and_display_table(visible_df)

            # Dates + Delay (once per day)
            recorded_str, effective_str, delay_str = (
                self.data.get_delay_metadata_for_day(selected_lei, selected_date)
            )

            parts = [
                f"**RecordedAt:** {recorded_str}",
                f"**EffectiveDate:** {effective_str}",
                f"**Delay:** {delay_str}",
            ]

            display(Markdown("<br/>" + " &nbsp;&nbsp; ".join(parts)))
            display(
                Markdown(
                    "<em>RecordedAt</em> is when the change was captured in the system, "
                    "whereas <em>EffectiveDate</em> is when the change became legally "
                    "effective.<br>"
                )
            )

    # ---------- Public API ----------

    def display(self) -> None:
        """
        Build and show widgets, using the attached data provider.
        """
        # Ensure data is present
        if not getattr(self.data, "lei_data_dict", {}):
            self.data.fetch_data()

        self.description_widget = widgets.HTML(
            "Retrieve the previous value that was used before a given "
            "LegalEntityEvent was triggered for a selected LEI and date."
        )

        self.lei_dropdown = widgets.Dropdown(
            options=self.data.leis,
            description="LEI:",
            layout=widgets.Layout(width="350px"),
            style={"description_width": "130px"},
        )

        self.date_dropdown = widgets.Dropdown(
            options=[],
            description="Modification Date:",
            layout=widgets.Layout(width="350px"),
            style={"description_width": "130px"},
        )

        self.output_widget = widgets.Output()

        # Link callbacks
        self.lei_dropdown.observe(self._update_dates, names="value")
        self.date_dropdown.observe(self._update_table_multi, names="value")

        display(self.description_widget)
        display(self.lei_dropdown)
        display(self.date_dropdown)
        display(self.output_widget)

        # Initialize UI
        if self.data.leis:
            self._update_dates({"new": self.data.leis[0]})

class NameAddressResultVisualizer:
    """Format LEI-to-ISO20022 result tables for notebook and Excel output."""

    def __init__(
        self,
        text_formatter=None,
        special_text_columns=None,
        special_text_passthrough_values=None,
        highlight_fill_hex: str = "DFFBC0",
        highlight_css: str = "#DFFBC0",
        default_bg_css: str = "#FFFFFF",
        default_text_css: str = "#000000",
        column_width_px: int = 260,
        column_width_px_map: Optional[dict[str, int]] = None,
        column_bg_css_map: Optional[dict[str, str]] = None,
        column_width_excel: int = 36,
    ) -> None:
        self.text_formatter = text_formatter
        self.special_text_columns = special_text_columns or []
        self.special_text_passthrough_values = set(special_text_passthrough_values or [])
        self.highlight_fill_hex = highlight_fill_hex
        self.highlight_css = highlight_css
        self.default_bg_css = default_bg_css
        self.default_text_css = default_text_css
        self.column_width_px = column_width_px
        self.column_width_px_map = column_width_px_map or {}
        self.column_bg_css_map = column_bg_css_map or {}
        self.column_width_excel = column_width_excel

    @staticmethod
    def _normalize_value(value):
        """Normalize values used as mapping keys."""
        if value is None:
            return ""
        return str(value).strip().lower()

    def _prepare_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        prepared = df.copy()

        if self.text_formatter:
            for col in self.special_text_columns:
                if col in prepared.columns:
                    prepared[col] = prepared[col].apply(
                        lambda x: x
                        if x in self.special_text_passthrough_values
                        else self.text_formatter(x)
                    )

        return prepared

    @staticmethod
    def _as_pre(value):
        """Render values in a `<pre>` block for notebook display."""
        if pd.isna(value) or value is None:
            return ""
        return (
            "<pre style='margin:0; white-space:pre-wrap; "
            f"font-family:monospace;'>{str(value)}</pre>"
        )

    def _highlight_selected_cells(
        self,
        row,
        source_column_to_target_map,
    ):
        styles = pd.Series("", index=row.index)

        for source_col, mapping in source_column_to_target_map.items():
            selected_value = self._normalize_value(row.get(source_col, ""))
            target_col = mapping.get(selected_value)

            if target_col and str(row.get(target_col, "")).strip():
                styles[target_col] = f"background-color: {self.highlight_css} !important;"

        return styles

    def display_notebook(
        self,
        result_df: pd.DataFrame,
        visible_columns,
        helper_columns=None,
        source_column_to_target_map=None,
        preformatted_columns=None,
    ):
        """Return a styled dataframe for notebook display.

        ``source_column_to_target_map`` maps source selector values to
        destination columns that should be highlighted.
        """
        helper_columns = helper_columns or []
        source_column_to_target_map = source_column_to_target_map or {}
        preformatted_columns = preformatted_columns or []

        df = self._prepare_dataframe(result_df)
        style_df = df[list(visible_columns) + list(helper_columns)].copy()

        for col in preformatted_columns:
            if col in style_df.columns:
                style_df[col] = style_df[col].apply(self._as_pre)

        styler = style_df.style

        if source_column_to_target_map:
            styler = styler.apply(
                lambda row: self._highlight_selected_cells(row, source_column_to_target_map),
                axis=1,
            )

        if helper_columns:
            styler = styler.hide(axis="columns", subset=helper_columns)

        if preformatted_columns:
            styler = styler.format(
                {col: (lambda x: x) for col in preformatted_columns},
                escape=None,
            )

        styler = styler.set_properties(
            **{
                "text-align": "left",
                "vertical-align": "top",
                "max-width": f"{self.column_width_px}px",
                "min-width": f"{self.column_width_px}px",
                "width": f"{self.column_width_px}px",
                "white-space": "pre-wrap",
                "color": self.default_text_css,
            }
        ).set_table_styles(
            [
                {
                    "selector": "table",
                    "props": [
                        ("border-collapse", "collapse"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [
                        ("background-color", self.default_bg_css),
                        ("border", "1px solid #cccccc !important"),
                    ],
                },
                {
                    "selector": "th",
                    "props": [
                        ("text-align", "left"),
                        ("vertical-align", "top"),
                        ("max-width", f"{self.column_width_px}px"),
                        ("min-width", f"{self.column_width_px}px"),
                        ("width", f"{self.column_width_px}px"),
                        ("background-color", self.default_bg_css),
                        ("color", self.default_text_css),
                        ("border", "1px solid #cccccc !important"),
                    ],
                }
            ]
        )

        if self.column_width_px_map:
            # Use CSS selectors to target specific columns
            col_index_map = {col: idx for idx, col in enumerate(style_df.columns)}
            width_styles = []
            for col, width in self.column_width_px_map.items():
                if col not in col_index_map:
                    continue
                width_css = f"{int(width)}px"
                col_idx = col_index_map[col]
                width_styles.extend(
                    [
                        {
                            "selector": f"th.col_heading.level0.col{col_idx}",
                            "props": [
                                ("max-width", width_css),
                                ("min-width", width_css),
                                ("width", width_css),
                            ],
                        },
                        {
                            "selector": f"td.col{col_idx}",
                            "props": [
                                ("max-width", width_css),
                                ("min-width", width_css),
                                ("width", width_css),
                            ],
                        },
                    ]
                )
            if width_styles:
                styler = styler.set_table_styles(width_styles, overwrite=False)

        if self.column_bg_css_map:
            col_index_map = {col: idx for idx, col in enumerate(style_df.columns)}
            bg_styles = []
            for col, color in self.column_bg_css_map.items():
                if col not in col_index_map or not color:
                    continue
                col_idx = col_index_map[col]
                bg_styles.extend(
                    [
                        {
                            "selector": f"th.col_heading.level0.col{col_idx}",
                            "props": [("background-color", color)],
                        },
                        {
                            "selector": f"td.col{col_idx}",
                            "props": [("background-color", color)],
                        },
                    ]
                )
            if bg_styles:
                styler = styler.set_table_styles(bg_styles, overwrite=False)

        return styler

    def save_excel(
        self,
        result_df: pd.DataFrame,
        output_file: str,
        visible_columns,
        source_column_to_target_map=None,
    ) -> None:
        """Save selected columns to Excel and apply optional highlights."""
        source_column_to_target_map = source_column_to_target_map or {}

        df = self._prepare_dataframe(result_df)
        export_df = df[list(visible_columns)].copy()
        export_df.to_excel(output_file, index=False)

        wb = load_workbook(output_file)
        ws = wb.active

        highlight_fill = PatternFill(
            fill_type="solid",
            start_color=self.highlight_fill_hex,
            end_color=self.highlight_fill_hex,
        )

        col_idx = {col: i + 1 for i, col in enumerate(visible_columns)}

        for column_cells in ws.columns:
            col_letter = column_cells[0].column_letter
            ws.column_dimensions[col_letter].width = self.column_width_excel

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                    horizontal="left",
                )

        if source_column_to_target_map:
            # enumerate over displayed/exported rows, not DataFrame index labels
            for excel_row, row in enumerate(df.to_dict("records"), start=2):
                for source_col, mapping in source_column_to_target_map.items():
                    selected_value = self._normalize_value(row.get(source_col, ""))
                    target_col = mapping.get(selected_value)

                    if (
                        target_col
                        and target_col in col_idx
                        and str(row.get(target_col, "")).strip()
                    ):
                        ws.cell(row=excel_row, column=col_idx[target_col]).fill = highlight_fill

        wb.save(output_file)